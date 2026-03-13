from __future__ import annotations

import base64
import csv
import io
from pathlib import Path
import re

from openai import OpenAI
from openpyxl import load_workbook
from PIL import Image
from pypdf import PdfReader

from app.core.config import Settings, get_settings

TEXT_SUFFIXES = {".txt", ".md"}
CSV_SUFFIXES = {".csv", ".tsv"}
SPREADSHEET_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
BINARY_PREFIXES = ("%PDF-", "PK\x03\x04")
OCR_IMAGE_MIN_AREA = 50_000
OCR_MAX_IMAGES_PER_PAGE = 3
OCR_MAX_COMPLETION_TOKENS = 1_800
OCR_PROMPT = (
    "Extract all visible Korean and English text from these page images in natural reading order. "
    "Preserve headings, bullet markers, table rows, and line breaks where reasonable. "
    "Do not summarize, interpret, or explain. Return plain text only."
)


def looks_like_binary_text(text: str | None) -> bool:
    if not text:
        return False
    sample = text.lstrip()[:256]
    if any(sample.startswith(prefix) for prefix in BINARY_PREFIXES):
        return True
    return any((ord(char) < 32 and char not in "\n\r\t") for char in sample)


def _read_text_file(path: Path) -> str:
    encodings = ("utf-8-sig", "utf-8", "cp949", "euc-kr", "utf-16")
    for encoding in encodings:
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
        except Exception:
            break
    return path.read_text(encoding="utf-8", errors="ignore")


def _extract_text_pages(path: Path) -> list[tuple[int | None, str]]:
    text = _read_text_file(path).strip()
    if not text:
        return []
    return [(1, text)]


def _extract_csv_pages(path: Path) -> list[tuple[int | None, str]]:
    raw_text = _read_text_file(path)
    if not raw_text.strip():
        return []

    sample = raw_text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except Exception:
        dialect = csv.excel_tab if path.suffix.lower() == ".tsv" else csv.excel

    reader = csv.reader(io.StringIO(raw_text), dialect=dialect)
    lines: list[str] = []
    for row in reader:
        values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
        if values:
            lines.append(" | ".join(values))
    if not lines:
        return []
    return [(1, "\n".join(lines))]


def _extract_spreadsheet_pages(path: Path) -> list[tuple[int | None, str]]:
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        pages: list[tuple[int | None, str]] = []
        for index, sheet in enumerate(workbook.worksheets, start=1):
            lines = [f"[Sheet] {sheet.title}"]
            for row in sheet.iter_rows(values_only=True):
                values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if values:
                    lines.append(" | ".join(values))
            if len(lines) > 1:
                pages.append((index, "\n".join(lines)))
        return pages
    finally:
        workbook.close()


def _build_ocr_client(settings: Settings) -> OpenAI | None:
    if not settings.ocr_enabled or not settings.openai_api_key:
        return None
    return OpenAI(
        api_key=settings.openai_api_key,
        base_url=settings.openai_base_url,
        timeout=settings.openai_timeout_seconds,
    )


def _select_page_images(page) -> list[Image.Image]:
    images: list[tuple[int, Image.Image]] = []
    try:
        for page_image in page.images:
            image = getattr(page_image, "image", None)
            if image is None:
                continue
            area = image.width * image.height
            if area < OCR_IMAGE_MIN_AREA:
                continue
            images.append((area, image))
    except Exception:
        return []

    if not images:
        return []
    images.sort(key=lambda item: item[0], reverse=True)
    max_area = images[0][0]
    return [image for area, image in images if area >= max_area * 0.25][:OCR_MAX_IMAGES_PER_PAGE]


def _image_to_data_url(image: Image.Image) -> str:
    buffer = io.BytesIO()
    image.save(buffer, format="PNG")
    encoded = base64.b64encode(buffer.getvalue()).decode("ascii")
    return f"data:image/png;base64,{encoded}"


def _extract_chat_message_text(content: object) -> str:
    if isinstance(content, str):
        return content.strip()
    if not isinstance(content, list):
        return ""

    fragments: list[str] = []
    for item in content:
        text = None
        if isinstance(item, dict):
            text = item.get("text")
        else:
            text = getattr(item, "text", None)
        if isinstance(text, str) and text.strip():
            fragments.append(text.strip())
    return "\n".join(fragments).strip()


def _ocr_pdf_page(page, client: OpenAI, *, model: str) -> str:
    images = _select_page_images(page)
    if not images:
        return ""

    content: list[dict[str, object]] = [{"type": "text", "text": OCR_PROMPT}]
    for image in images:
        content.append(
            {
                "type": "image_url",
                "image_url": {"url": _image_to_data_url(image)},
            }
        )

    try:
        response = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": content}],
            temperature=0,
            max_completion_tokens=OCR_MAX_COMPLETION_TOKENS,
        )
    except Exception:
        return ""

    if not response.choices:
        return ""
    return _extract_chat_message_text(response.choices[0].message.content)


def _extract_pdf_pages(path: Path) -> list[tuple[int | None, str]]:
    try:
        reader = PdfReader(str(path))
    except Exception:
        return []

    settings = get_settings()
    ocr_client = _build_ocr_client(settings)
    pages: list[tuple[int | None, str]] = []
    for index, page in enumerate(reader.pages, start=1):
        try:
            text = (page.extract_text() or "").strip()
        except Exception:
            text = ""

        if text:
            pages.append((index, text))
            continue
        if ocr_client is None:
            continue

        ocr_text = _ocr_pdf_page(page, ocr_client, model=settings.openai_model_extraction).strip()
        if ocr_text:
            pages.append((index, ocr_text))
    return pages


def extract_text_pages_from_path(path: Path) -> list[tuple[int | None, str]]:
    suffix = path.suffix.lower()
    if suffix in TEXT_SUFFIXES:
        return _extract_text_pages(path)
    if suffix in CSV_SUFFIXES:
        return _extract_csv_pages(path)
    if suffix in SPREADSHEET_SUFFIXES:
        return _extract_spreadsheet_pages(path)
    if suffix == ".pdf":
        return _extract_pdf_pages(path)
    return []


def extract_text_from_path(path: Path) -> tuple[str, bool]:
    pages = extract_text_pages_from_path(path)
    text = "\n\n".join(page_text for _, page_text in pages if page_text).strip()
    ocr_required = path.suffix.lower() == ".pdf" and len(text.strip()) < 120
    return text, ocr_required
